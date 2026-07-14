import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Read JSON file
with open(r'd:\Project\automationexercise\src\test\resources\testdata\test-data.json', 'r') as f:
    data = json.load(f)

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Style definitions
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def add_styled_header(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

# Sheet 1: Accounts
ws = wb.create_sheet("Accounts")
add_styled_header(ws, 1, ["Test Case", "Field", "Value"])
row = 2
for test_case, data_dict in data['accounts'].items():
    if isinstance(data_dict, dict):
        for field, value in data_dict.items():
            if isinstance(value, dict):
                for sub_field, sub_value in value.items():
                    ws.cell(row=row, column=1).value = test_case
                    ws.cell(row=row, column=2).value = f"{field}_{sub_field}"
                    ws.cell(row=row, column=3).value = str(sub_value)
                    row += 1
            else:
                ws.cell(row=row, column=1).value = test_case
                ws.cell(row=row, column=2).value = field
                ws.cell(row=row, column=3).value = str(value)
                row += 1

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 40

# Sheet 2: Products
ws = wb.create_sheet("Products")
add_styled_header(ws, 1, ["Category", "Field", "Value"])
row = 2
for cat, cat_data in data['product'].items():
    if isinstance(cat_data, dict):
        for field, value in cat_data.items():
            ws.cell(row=row, column=1).value = cat
            ws.cell(row=row, column=2).value = field
            ws.cell(row=row, column=3).value = str(value)
            row += 1
    else:
        ws.cell(row=row, column=1).value = cat
        ws.cell(row=row, column=2).value = "value"
        ws.cell(row=row, column=3).value = str(cat_data)
        row += 1

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 40

# Sheet 3: Contact Us
ws = wb.create_sheet("ContactUs")
add_styled_header(ws, 1, ["Field", "Value"])
row = 2
for field, value in data['contactUs'].items():
    ws.cell(row=row, column=1).value = field
    ws.cell(row=row, column=2).value = str(value)
    row += 1

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 40

# Sheet 4: Category
ws = wb.create_sheet("Category")
add_styled_header(ws, 1, ["Field", "Value"])
row = 2
for field, value in data['category'].items():
    ws.cell(row=row, column=1).value = field
    ws.cell(row=row, column=2).value = str(value)
    row += 1

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 40

# Sheet 5: Cart
ws = wb.create_sheet("Cart")
add_styled_header(ws, 1, ["Field", "Value"])
row = 2
for field, value in data['cart'].items():
    ws.cell(row=row, column=1).value = field
    ws.cell(row=row, column=2).value = str(value)
    row += 1

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 40

# Sheet 6: Messages
ws = wb.create_sheet("Messages")
add_styled_header(ws, 1, ["Field", "Value"])
row = 2
for field, value in data['messages'].items():
    ws.cell(row=row, column=1).value = field
    ws.cell(row=row, column=2).value = str(value)
    row += 1

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 50

# Sheet 7: Home
ws = wb.create_sheet("Home")
add_styled_header(ws, 1, ["Field", "Value"])
row = 2
for field, value in data['home'].items():
    if isinstance(value, dict):
        for sub_field, sub_value in value.items():
            ws.cell(row=row, column=1).value = f"{field}_{sub_field}"
            ws.cell(row=row, column=2).value = str(sub_value)
            row += 1
    else:
        ws.cell(row=row, column=1).value = field
        ws.cell(row=row, column=2).value = str(value)
        row += 1

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 50

# Save file
output_path = r'd:\Project\automationexercise\src\test\resources\testdata\test-data.xlsx'
wb.save(output_path)
print(f"[OK] Excel file created successfully: {output_path}")
print(f"[OK] Total sheets: {len(wb.sheetnames)}")
print(f"[OK] Sheets: {', '.join(wb.sheetnames)}")
